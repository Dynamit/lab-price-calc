#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Regenerate the site's JSON data files from the source Excel workbooks.

The site (index.html / script.js) reads three JSON files from assets/data/:
  - lab_prices.json         : list of {test_code, test_name, prices:{private,tourist}}
  - lab_details.json        : Ramat HaHayal test book (6 fields per test)
  - lab_details_haifa.json  : Haifa test book (10 fields per test)

All three are GENERATED from the Excel files listed in CONFIG below (the .xlsx
files are gitignored — only the generated JSON is committed). When a new price
list or test book arrives:
  1. Drop the new .xlsx into the project root.
  2. Update the matching filename in CONFIG below.
  3. Run:  python tools/build_data.py
  4. Review `git diff` on assets/data/*.json, then commit.

Tests are keyed by "קוד תפנית" so prices and details line up. Prices are identical
for both branches; the branch selector only changes which details book is shown.

Requires: openpyxl  (pip install openpyxl)
"""

import json
import os
import re
import sys

import openpyxl

# Project root = parent of this tools/ directory.
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "assets", "data")

CONFIG = {
    # Price list: columns A-D = code, name, private price, tourist price.
    "prices_xlsx": "מעבדה 05.xlsx",
    # Ramat HaHayal test book: sheet 0, header row 4, data from row 5.
    "ramat_hahayal_xlsx": "ספר בדיקות- אסותא רמת החייל 01.26 (2).xlsx",
    # Haifa test book: sheet 0, header row 12, data from row 13.
    "haifa_xlsx": "מדריך בדיקות אסותא חיפה 27052026.xlsx",
}


def clean(value):
    """Trim a cell to a string ('' for empty)."""
    return "" if value is None else str(value).strip()


def clean_ml(value):
    """Like clean() but normalize CRLF -> LF (Haifa cells use multi-line text)."""
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def build_prices(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    prices = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        prices.append({
            "test_code": clean(r[0]),
            "test_name": clean(r[1]),
            "prices": {
                "private": int(round(r[2])) if r[2] is not None else 0,
                "tourist": int(round(r[3])) if r[3] is not None else 0,
            },
        })
    wb.close()
    return prices


# Ramat HaHayal: JSON field -> 1-based column number (sheet 0, header row 4).
RH_FIELDS = {
    "patient_preparation_conditions": 10,
    "tubes": 13,
    "sampling_conditions": 14,
    "transport_instructions": 15,
    "execution_time_info": 16,
    "results_time": 17,
}


def build_ramat_hahayal(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    details = {}
    for r in ws.iter_rows(min_row=5, values_only=True):
        if all(c is None for c in r):
            continue
        # col C = קוד תפנית. Some cells are multi-line ("4540\nאיסוף שתן-45201") — key by the
        # first line only so the detail matches its numeric price code (otherwise it's orphaned).
        code = clean(r[2]).split("\n")[0].strip()
        if not code or code in details:  # first occurrence wins
            continue
        details[code] = {k: clean(r[i - 1]) for k, i in RH_FIELDS.items()}
    wb.close()
    return details


# Haifa: JSON field -> 1-based column number (sheet 0, header row 12).
# Columns requested for display: C,D,H,I,J,M,O,P,Q,R.
HAIFA_FIELDS = {
    "test_name_book": 3,                    # C
    # code_tfnit (D) is set to the canonical code below
    "patient_preparation_conditions": 8,    # H
    "sampling_conditions": 9,               # I  (תנאי לקיחה)
    "tubes": 10,                            # J
    "execution_time_info": 13,              # M
    "performing_lab": 15,                   # O
    "results_time": 16,                     # P
    "storage_conditions": 17,               # Q  (תנאי שימור)
    "transport_instructions": 18,           # R
}


def build_haifa(path, price_codes):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    details, dups, no_price, bad_code = {}, [], 0, 0
    for r in ws.iter_rows(min_row=13, values_only=True):
        if all(c is None for c in r):
            continue
        raw = clean_ml(r[3])  # col D = קוד תפנית
        code = raw.split("\n")[0].strip() if raw else ""
        if not re.fullmatch(r"\d+", code):
            bad_code += 1
            continue
        if code not in price_codes:
            no_price += 1
            continue
        if code in details:
            dups.append(code)
            continue
        entry = {k: clean_ml(r[i - 1]) for k, i in HAIFA_FIELDS.items()}
        entry["code_tfnit"] = code  # D
        details[code] = entry
    wb.close()
    return details, sorted(set(dups)), no_price, bad_code


def write_json(name, data, sort_keys=False):
    path = os.path.join(DATA_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=sort_keys)
        f.write("\n")
    return path


def main():
    paths = {k: os.path.join(ROOT, v) for k, v in CONFIG.items()}
    missing = [v for k, v in CONFIG.items() if not os.path.exists(paths[k])]
    if missing:
        print("ERROR: missing Excel file(s):", *missing, sep="\n  ")
        print("\nUpdate CONFIG in tools/build_data.py to match the files in the project root.")
        sys.exit(1)

    prices = build_prices(paths["prices_xlsx"])
    price_codes = set(p["test_code"] for p in prices)
    write_json("lab_prices.json", prices)

    rh = build_ramat_hahayal(paths["ramat_hahayal_xlsx"])
    write_json("lab_details.json", rh)

    haifa, dups, no_price, bad_code = build_haifa(paths["haifa_xlsx"], price_codes)
    write_json("lab_details_haifa.json", haifa, sort_keys=True)

    rh_match = len(price_codes & set(rh))
    hf_match = len(price_codes & set(haifa))
    print("Generated assets/data/*.json")
    print(f"  prices:            {len(prices)} tests ({len(price_codes)} unique codes)")
    print(f"  Ramat HaHayal:     {len(rh)} details  ({rh_match} matched to a price)")
    print(f"  Haifa:             {len(haifa)} details  ({hf_match} matched to a price)")
    print(f"  Haifa dropped:     {no_price} no-price, {bad_code} non-numeric code")
    if dups:
        print(f"  Haifa dup codes (first-wins): {', '.join(dups)}")
    print(f"  Coverage: {len(price_codes & (set(rh) | set(haifa)))} of {len(price_codes)} "
          f"tests have details in at least one branch.")


if __name__ == "__main__":
    main()
